import openpyxl as xl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side


class BasicExcelDecorator:

    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)

    def __init__(self, excel_path, sheet_name, table_size, upper_left=[1, 1]):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.row_headline, self.col_headline = upper_left
        self.row_length, self.col_length = table_size
        self.load_workbook()

    # wbを読み込む関数
    def load_workbook(self):
        self.wb = xl.load_workbook(self.excel_path)
        self.ws = self.wb[self.sheet_name]

    # wbを保存する関数
    def save(self):
        self.wb.save(self.excel_path)
        self.wb.close()

    # cellを中心に配置する
    def center_cell(self, cell):
        cell.alignment = Alignment(
                                    horizontal='center',
                                    vertical='center',
                                    wrap_text=False
                                    )
        return cell

    # 表の左上に表題を記述する関数
    def write_title(self, name='', font=Font(bold=True, color='ffffff')):
        self.ws.cell(column=self.col_headline, row=self.row_headline).value = name
        self.ws.cell(column=self.col_headline, row=self.row_headline).font = font

    # 装飾範囲内のcellかどうかの判定
    def is_in_range(self, cell):
        return self.row_headline <= cell.row <= self.row_headline + self.row_length and \
            self.col_headline <= cell.column <= self.col_headline + self.col_length

    # 背景の色付けを行う
    def coloring_background(self, cell, fill_color='ffffff', font_color='000000'):
        self.ws[cell.coordinate].fill = PatternFill(patternType='solid', fgColor=fill_color)
        self.ws[cell.coordinate].font = Font(color=font_color)

    # 見出しを指定色で色付けする
    def coloring_headline(self, cell, fill_color='000000', font_color='ffffff', color_mapping={}):
        if self.is_headline_cell(cell):
            self.ws[cell.coordinate].fill = PatternFill(patternType='solid', fgColor=fill_color)
            self.ws[cell.coordinate].font = Font(color=font_color)
            for color, value in color_mapping.items():
                if self.ws[cell.coordinate].value == value:
                    self.ws[cell.coordinate].fill = PatternFill(patternType='solid', fgColor=color)
                    self.ws[cell.coordinate].font = Font(color='000000')

    # cellが表の見出しかどうか判定する関数
    def is_headline_cell(self, cell):
        return (int(cell.column) == self.col_headline) or (int(cell.row) == self.row_headline)

    # 入力したcellの列名を返す関数
    def col_name(self, cell):
        return self.ws.cell(column=cell.column, row=self.row_headline).value

    # 入力したcellの行名を返す関数
    def row_name(self, cell):
        return self.ws.cell(column=self.col_headline, row=cell.row).value

    # 入力した列名のindexを返す関数
    def col_headline_index(self, col_name):
        for i in range(self.col_length):
            if self.ws.cell(column=self.col_headline+i, row=self.row_headline).value == col_name:
                return i+1
        raise KeyError('{} does not exist in the column'.format(col_name))

    # 入力した行名のindexを返す関数
    def row_headline_index(self, row_name):
        for i in range(self.row_length):
            if self.ws.cell(column=self.col_headline, row=self.row_headline+i).value == row_name:
                return i+1
        raise KeyError('{} does not exist in the row'.format(row_name))

    # 対象の列名を持つcellのフォーマットを%に変更する関数
    def to_percent(self, cell, pct_list):
        # セルのフォーマットをパーセンテージにする
        if pct_list and self.col_name(cell) in pct_list:
            cell.number_format = '0.00%'
        return cell
