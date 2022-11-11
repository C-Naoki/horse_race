import os
import openpyxl as xl
import pandas as pd


class BasicExcelWriter:
    def _append_df2xlsx(self, df, excel_path, sheet_name):
        self.wb = xl.load_workbook(excel_path)
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        writer.book = self.wb
        writer.sheets = {ws.title: ws for ws in self.wb.worksheets}
        startrow = writer.sheets[sheet_name].max_row+1
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, startcol=0)
        writer.save()

        self.ws = self.wb[sheet_name]
        self.row_length, self.col_length = df.shape
        self.col_heading = writer.sheets[sheet_name].max_row - self.row_length

    def _overwrite_df2xlsx(self, df, excel_path, sheet_name):
        writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
        df.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0)
        writer.save()
        self.wb = xl.load_workbook(excel_path)

        self.ws = self.wb[sheet_name]
        self.row_length, self.col_length = df.shape
        self.col_heading = writer.sheets[sheet_name].max_row - self.row_length

    def _mkdir(self, path):
        if not os.path.isdir(path):
            os.makedirs(path)

    def _mkxlsx(self, path):
        if not os.path.isfile(path):
            wb = xl.Workbook()
            wb.save(path)
