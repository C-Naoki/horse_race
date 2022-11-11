import re
from backend.module.record import BasicExcelDecorator
from openpyxl.styles import PatternFill


class ExcelDecorator(BasicExcelDecorator):

    rank_color_mapping = {
        'A': 'ffbf7f',
        'B': 'a8d3ff',
        'C': 'd3d3d3',
        'ABC': '50C878',
    }

    order_color_mapping = {
        1: 'ffbf7f',
        2: 'a8d3ff',
        3: 'd3d3d3',
        0: '50C878',
    }

    def __init__(self, excel_path, sheet_name, table_size, upper_left=[1, 1]):
        super().__init__(excel_path, sheet_name, table_size, upper_left)
        try:
            self.result_col_idx = self.col_headline_index('結果')
        except KeyError:
            pass

    def __call__(
        self,
        headline_color_mapping={},
        pct_list=[],
    ):
        for col in self.ws.columns:
            max_length = 0
            for cell in col:
                max_length = max(max_length, len(str(cell.value)))
                if not self.is_in_range(cell):
                    continue
                coord = cell.coordinate
                cell = self.center_cell(cell)
                self.coloring_background(cell)
                self.coloring_headline(cell, color_mapping=headline_color_mapping)
                self.coloring_table(cell)
                cell = self.to_percent(cell, pct_list)
                self.ws[coord].border = self.border
            adjusted_width = max_length * 2.08
            self.ws.column_dimensions[cell.column_letter].width = adjusted_width
        self.write_title(self.sheet_name)
        self.save()

    # カラムに対応した色付けを行う
    def coloring_table(self, cell):
        col_name = self.col_name(cell)
        if '本命馬ランク' == col_name:
            self.coloring_fav_rank(cell)
        elif '1着予想◎' == col_name:
            self.coloring_fukusho(cell, 'a8d3ff')
            self.coloring_tansho(cell)
        elif '2着予想○' == col_name or '3着予想▲' == col_name:
            self.coloring_fukusho(cell, color='a8d3ff')
        elif '4着予想△' == col_name or '5着予想☆' == col_name or '6着予想×' == col_name:
            self.coloring_fukusho(cell, color='d3d3d3')

    def coloring_fav_rank(self, cell):
        for rank, color in self.rank_color_mapping.items():
            if cell.value == rank:
                self.ws[cell.coordinate].fill = PatternFill(patternType='solid', fgColor=color)
                break

    def coloring_tansho(self, cell):
        top_3 = self.ws.cell(column=self.result_col_idx, row=cell.row).value.split(' → ')
        if top_3[0] == re.sub(r"\D", "", str(cell.value)):
            self.ws[cell.coordinate].fill = PatternFill(patternType='solid', fgColor='ffbf7f')

    def coloring_fukusho(self, cell, color):
        top_3 = self.ws.cell(column=self.result_col_idx, row=cell.row).value.split(' → ')
        for within_horse in top_3:
            if within_horse == re.sub(r"\D", "", str(cell.value)):
                self.ws[cell.coordinate].fill = PatternFill(patternType='solid', fgColor=color)
